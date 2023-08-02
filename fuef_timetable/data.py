from openpyxl import load_workbook

import datetime
from typing import List
from nocodb.nocodb import NocoDBProject, APIToken, JWTAuthToken
from nocodb.filters.raw_filter import RawFilter
from nocodb.infra.requests_client import NocoDBRequestsClient
from dotenv import dotenv_values

class Data:
    """Collection of entries."""

    def __init__(
        self,
        debug: bool = False
    ):

        env = dotenv_values('.env')

        # Usage with API Token
        client = NocoDBRequestsClient(
            # Your API Token retrieved from NocoDB conf
            APIToken(env["API-TOKEN"]),
            # Your nocodb root path
            env["URL"]
        )

        project = NocoDBProject(
            env["ORGANISATION"],
            env["PROJECT"]
        )

        table_name = env["TABLE"]

        table_rows = client.table_row_list(project, table_name)

#        wb = load_workbook(filename=path)
#        sheet = wb["Acts"]

        self.entries: List[Entry] = []
        if not debug:
            today_date = datetime.date.today()
        else:
            today_date = datetime.date(year=2023, month=8, day=3)

        for element in table_rows["list"]:
            if datetime.datetime.strftime(today_date, "%Y-%m-%d") in element["when"]:
                print(element)
                entry = Entry(element)
                self.entries.append(entry)

    def bar_stuff(self, debug: bool):
        """Mess with the best, die like the rest."""
        now = datetime.datetime.now()
        tmp = []
        for entry in self.entries:
            if entry.when >= now or entry.is_now_fn(debug):
                print(1)
                tmp.append(entry)
        return tmp[:2]


class Entry:
    """A single entry."""

    def __init__(self, element):
        if element["when"]:
            self.when = datetime.datetime.strptime(element["when"], "%Y-%m-%d %H:%M:%S%z")
        if not isinstance(self.when, datetime.datetime):
            self.is_valid = False
            return
        if element["artist"]:
            if not (element["artist"][0] is None):
                self.artist = element["artist"][0]
        if element["title"]:
            if not (element["title"][0] is None):
                self.title = element["title"][0]
        if element["location"]:
            if not (element["location"][0] is None):
                self.place = element["location"][0]["name"]
        if element["duration"]:
            if not(element["duration"][0] is None):
                self.duration = element["duration"][0]
                self.endtime = self.when + datetime.timedelta(0,int(self.duration))
        if element["genre"]:
            self.genre = element["genre"][0]

        #self.is_now = self.is_now_fn(debug)

        # print(ends)
        self.is_valid = True

    def is_now_fn(self, debug: bool) -> bool:
        """
        Determines wether an event is currently happening.
        Inkl. 10 Minutes before.
        """
        if self.duration != "–":
            duration = self.duration
        else:
            duration = 15
        if not debug:
            now = datetime.datetime.now()
        else:
            now = datetime.datetime(
                year=2022,
                month=7,
                day=28,
                hour=19,
                minute=25
            )
        starts = self.when - datetime.timedelta(minutes=10)
        ends = self.when + datetime.timedelta(minutes=duration)
        return now >= starts and now <= ends

    @staticmethod
    def get_value(sheet, row, column):
        """
        Gets the value from the sheet and replaces None with a dash.
        """
        value = sheet.cell(row=row, column=column).value
        if value is None:
            return "–"
        return value
